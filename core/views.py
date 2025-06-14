from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from .forms import CustomAuthenticationForm
from django.shortcuts import render
from django.db import connection
from django.urls import reverse
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
import requests
from django.http import JsonResponse
from django.db import connection, DataError, IntegrityError
from openpyxl import Workbook
from django.http import HttpResponse
import pytz
from django.utils import timezone
from datetime import datetime
from rest_framework import viewsets
from .serializers import *
from django.core.paginator import Paginator
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook
import openpyxl.utils
from django.contrib.auth import views as auth_views

# SERIALIZERS:
class RolUsuarioViewSet(viewsets.ModelViewSet):
    queryset = rolUsuario.objects.all()
    serializer_class = RolUsuarioSerializer

class UsuarioCustomViewSet(viewsets.ModelViewSet):
    queryset = usuarioCustom.objects.all()
    serializer_class = UsuarioCustomSerializer

class MarcaViewSet(viewsets.ModelViewSet):
    queryset = marca.objects.all()
    serializer_class = MarcaSerializer

class CategoriaProductoViewSet(viewsets.ModelViewSet):
    queryset = categoriaProducto.objects.all()
    serializer_class = CategoriaProductoSerializer

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = producto.objects.all()
    serializer_class = ProductoSerializer

class SucursalViewSet(viewsets.ModelViewSet):
    queryset = sucursal.objects.all()
    serializer_class = SucursalSerializer

class CarritoViewSet(viewsets.ModelViewSet):
    queryset = Carrito.objects.all()
    serializer_class = CarritoSerializer

class ItemCarritoViewSet(viewsets.ModelViewSet):
    queryset = ItemCarrito.objects.all()
    serializer_class = ItemCarritoSerializer

class SeguimientoViewSet(viewsets.ModelViewSet):
    queryset = Seguimiento.objects.all()
    serializer_class = SeguimientoSerializer

class PedidoViewSet(viewsets.ModelViewSet):
    queryset = Pedido.objects.all()
    serializer_class = PedidoSerializer

class ItemPedidoViewSet(viewsets.ModelViewSet):
    queryset = ItemPedido.objects.all()
    serializer_class = ItemPedidoSerializer


# VIEWS
def index(request):
	return render(request, 'core/index.html')

@login_required
def shop(request):
    
    productos = lista_productos()
    paginator = Paginator(productos, 4)

    page = request.GET.get('page', 1)
    productos = paginator.get_page(page)
    
    data = {
        'listado': productos,
        'MEDIA_URL': settings.MEDIA_URL,
        'paginator': paginator
    }
    return render(request, 'core/shop.html', data)

def about(request):
	return render(request, 'core/about.html')

def services(request):
	return render(request, 'core/services.html')

@login_required
def contact(request):
	return render(request, 'core/contact.html')

@login_required
def cart(request):
    try:
        usuario = request.user
        carrito = Carrito.objects.get(usuario=usuario)
        items = carrito.itemcarrito_set.all()
        
        api_mindicador = requests.get('https://mindicador.cl/api/')
        divisas = api_mindicador.json()
        tasa_dolar = divisas['dolar']['valor']

        subtotal = sum(item.precio_total() for item in items)
        subtotal_dolar = round(subtotal / tasa_dolar, 2)
        total = subtotal 
        total_dolar = subtotal_dolar

        data = {
            'carrito': carrito,
            'items': items,  
            'subtotal': subtotal,
            'total': total,
            'subtotal_dolar': subtotal_dolar,
            'total_dolar': total_dolar,
            'MEDIA_URL': settings.MEDIA_URL,
        }
        
        return render(request, 'core/cart.html', data)
    
    except Carrito.DoesNotExist:

        data = {
            'carrito': None,
            'items': None,
            'subtotal': 0,
            'total': 0,
            'subtotal_dolar': 0,
            'total_dolar': 0,
            'MEDIA_URL': settings.MEDIA_URL,
        }
        # Muestra un mensaje informativo
        messages.info(request, 'Tu carrito está vacío. Añade productos para continuar.')
        return render(request, 'core/cart.html', data)


@login_required
def checkout(request):
    try:
        usuario = request.user
        carrito = Carrito.objects.get(usuario=usuario)
        items = carrito.itemcarrito_set.all()
        
        sucursales = sucursal.objects.all()

        api_mindicador = requests.get('https://mindicador.cl/api/')
        divisas = api_mindicador.json()
        tasa_dolar = divisas['dolar']['valor']

        subtotal = sum(item.precio_total() for item in items)
        subtotal_dolar = round(subtotal / tasa_dolar, 2)
        total = subtotal 
        total_dolar = subtotal_dolar

        subtotal_dolar_str = '{:.2f}'.format(subtotal_dolar) 
        total_dolar_str = '{:.2f}'.format(total_dolar)

        if request.method == "POST":
            form = CheckoutForm(request.POST, request.FILES)
            if form.is_valid():
                metodo_pago = form.cleaned_data['metodo_pago']
                
                pedido = Pedido.objects.create(
                    carrito=carrito,
                    tipo_entrega=form.cleaned_data['tipo_entrega'],
                    direccion=form.cleaned_data['direccion'],
                    nombre=form.cleaned_data['nombre'],
                    apellido=form.cleaned_data['apellido'],
                    run=form.cleaned_data['run'],
                    correo=form.cleaned_data['correo'],
                    sucursal=form.cleaned_data['sucursal'],
                    comprobante_pago=form.cleaned_data['comprobante_pago']
                )

                if metodo_pago == 'transferencia':
                    pedido.comprobante_pago = form.cleaned_data['comprobante_pago']
                    pedido.estado = Seguimiento.objects.get(descripcion='Esperando comprobante de pago')
                    pedido.save()
                    return redirect('confirmacion_pedido')
        else:
            form = CheckoutForm()

        data = {
            'carrito': carrito,
            'items': items,
            'subtotal': subtotal,
            'total': total,
            'subtotal_dolar': subtotal_dolar_str,
            'total_dolar': total_dolar_str,
            'sucursales': sucursales,
            'MEDIA_URL': settings.MEDIA_URL,
            'form': form,
        }

        return render(request, 'core/checkout.html', data)
    except Exception as e:
        print("ERROR EN CHECKOUT: ", e)


@login_required
def thankyou(request):
	return render(request, 'core/thankyou.html')

def register(request):
    form = RegisterForm()
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            # Asignar el rol por defecto Cliente
            user.idRol = rolUsuario.objects.get(nombreRol='Cliente')
            
            # Guardar el correo_usuario en el campo email predeterminado
            user.email = user.correo_usuario

            user.save()
            return redirect("index")

    return render(request, 'registration/register.html', {'form': form})

class CustomLoginView(LoginView):
    authentication_form = CustomAuthenticationForm
    template_name = 'registration/register.html'
    success_url = reverse_lazy('core/index.html')

    def form_valid(self, form):
        user = form.get_user()
        login(self.request, user)
        return super().form_valid(form)

def handle_uploaded_file(f):
    with open('media/productos/' + f.name, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

# VISTAS CRUD PRODUCTOS (AÑADIR, ACTUALIZAR, ELIMINAR)
def addProduct(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto_instance = form.save(commit=False)
            
            imagenProducto = request.FILES['imagenProducto']
            with open('media/productos/' + imagenProducto.name, 'wb+') as destination:
                for chunk in imagenProducto.chunks():
                    destination.write(chunk)
            producto_instance.imagenProducto = 'productos/' + imagenProducto.name

            try:
                producto_instance.save()
                return redirect("shop")
            except Exception as e:
                error_msg = "Hubo un error al agregar el producto. Por favor, inténtalo de nuevo más tarde."
                print("ERROR EN AGREGAR PRODUCTO: ", e)
                return render(request, 'core/shop.html', {'form': form, 'error_msg': error_msg})
    else:
        form = ProductoForm()
    return render(request, 'core/addproduct.html', {'form': form})


# LISTAR PRODUCTOS (REEMPLAZO DE SP_GET_PRODUCTOS)
def lista_productos():
    return producto.objects.all()


# DETALLE DE PRODUCTO
def detalle_producto(request, idProducto):
    producto_instance = producto.objects.get(idProducto=idProducto)
    api_mindicador = requests.get('https://mindicador.cl/api/')
    divisas = api_mindicador.json()
    tasa_dolar = divisas['dolar']['valor']

    precio_dolares = round(producto_instance.precioProducto / tasa_dolar, 2)

    data = {
        'producto': producto_instance,
        'MEDIA_URL': settings.MEDIA_URL,
        'precio_dolares': precio_dolares,
    }
    return render(request, 'core/detalle_producto.html', data)


# MODIFICAR PRODUCTO
def modificar_producto(request, idProducto):
    producto_instance = get_object_or_404(producto, idProducto=idProducto)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto_instance)
        if form.is_valid():
            try:
                producto_modificado = form.save(commit=False)
                
                imagenProducto = request.FILES.get('imagenProducto', None)
                if imagenProducto:
                    with open('media/productos/' + imagenProducto.name, 'wb+') as destination:
                        for chunk in imagenProducto.chunks():
                            destination.write(chunk)
                    producto_modificado.imagenProducto = 'productos/' + imagenProducto.name
                
                producto_modificado.save()
                messages.success(request, 'Producto modificado exitosamente.')
                return redirect('detalle_producto', idProducto=idProducto)

            except DataError as e:
                messages.error(request, "Hubo un problema al ingresar los datos, revise nuevamente por favor.")
                print(e)
            except IntegrityError as e:
                messages.error(request, "Hubo un problema con la integridad de los datos. Por favor, inténtalo de nuevo.")
            except Exception as e:
                messages.error(request, "Ocurrió un error inesperado. Por favor, inténtalo de nuevo.")
    else:
        form = ProductoForm(instance=producto_instance)
    return render(request, 'core/modificar_producto.html', {'form': form})


# ELIMINAR PRODUCTO (REEMPLAZO DE SP_DELETE_PRODUCTO)
def eliminar_producto(request, idProducto):
    producto.objects.filter(idProducto=idProducto).delete()
    return redirect(to='shop')


# PANEL DE ADMINISTRACIÓN
def panel_administracion(request):
    return render(request, 'core/panel_administracion.html') 


# GESTIÓN DE USUARIOS (PENDIENTE: lista_usuarios)
def gestion_usuarios(request):
    data = {
        'usuarios': lista_usuarios() 
    }
    return render(request, 'core/gestion_usuarios.html', data)


# GESTIÓN DE USUARIOS
def lista_usuarios():
    return usuarioCustom.objects.all().order_by('-id')


def eliminar_usuario(request, id):
    try:
        usuarioCustom.objects.filter(id=id).delete()
        messages.success(request, 'Usuario eliminado correctamente.')
    except Exception as e:
        print("Error en eliminar usuario: ", e)
        messages.error(request, 'Error al eliminar usuario.')
    return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})


def modificar_usuario(request, p_id):
    usuario_instance = get_object_or_404(usuarioCustom, id=p_id)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario_instance)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Usuario modificado correctamente!')
            except Exception as e:
                print("ERROR al modificar usuario:", e)
                messages.error(request, "Ocurrió un error al modificar el usuario.")
            return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})
    else:
        form = UsuarioForm(instance=usuario_instance)
    return render(request, 'core/modificar_usuario.html', {'form': form})


def post_usuario(p_username, p_run, p_pnombre, p_ap_paterno, p_correo_usuario, p_fecha_nacimiento, p_direccion, p_idRol, p_password):
    try:
        password_encrypted = make_password(p_password)
        usuarioCustom.objects.create(
            username=p_username,
            run=p_run,
            pnombre=p_pnombre,
            ap_paterno=p_ap_paterno,
            correo_usuario=p_correo_usuario,
            fecha_nacimiento=p_fecha_nacimiento,
            direccion=p_direccion,
            idRol_id=p_idRol,
            password=password_encrypted
        )
        return print("Usuario creado correctamente")
    except Exception as e:
        print("ERROR EN POST USUARIO: ", e)


def add_user(request):
    if request.method == 'POST':
        form = RegisterUserAdminForm(request.POST)
        if form.is_valid():
            form.save()  # Crea en el modelo de auth
            p_username = form.cleaned_data['username']
            p_run = form.cleaned_data['run']
            p_pnombre = form.cleaned_data['pnombre']
            p_ap_paterno = form.cleaned_data['ap_paterno']
            p_correo_usuario = form.cleaned_data['correo_usuario']
            p_fecha_nacimiento = form.cleaned_data['fecha_nacimiento']
            p_direccion = form.cleaned_data['direccion']
            p_idRol = form.cleaned_data['idRol'].pk
            p_password = form.cleaned_data['password1']

            post_usuario(
                p_username,
                p_run,
                p_pnombre,
                p_ap_paterno,
                p_correo_usuario,
                p_fecha_nacimiento,
                p_direccion,
                p_idRol,
                p_password
            )

            messages.success(request, 'Usuario agregado correctamente!')
            return render(request, 'core/gestion_usuarios.html', {'usuarios': lista_usuarios()})
    else:
        form = RegisterUserAdminForm()
    return render(request, 'core/addusuario.html', {'form': form})



# CARRITO DE COMPRAS

def agregar_al_carrito(request, idProducto):
    producto_cart = get_object_or_404(producto, pk=idProducto)
    carrito, created = Carrito.objects.get_or_create(usuario=request.user)
    item, item_created = ItemCarrito.objects.get_or_create(carrito=carrito, producto=producto_cart)

    if not item_created:
        item.cantidad += 1
        item.save()
    else:
        item.cantidad = 1
        item.save()

    # Disminuir el stock del producto
    producto_cart.disminuir_stock(1)

    return redirect(to="cart")


def eliminar_del_carrito(request, itemcarrito_id):
    item = get_object_or_404(ItemCarrito, pk=itemcarrito_id, carrito__usuario=request.user)
    producto = item.producto

    # Incrementar la cantidad disponible en el stock del producto
    producto.stockProducto += item.cantidad
    producto.save()

    item.delete()

    return redirect('cart')


# CREAR PEDIDO
def crear_pedido(request):
    if request.method == 'POST':
        usuario = request.user
        carrito = usuario.carrito

        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        run = request.POST.get('run') 
        tipo_entrega = request.POST.get('tipo_entrega')
        sucursal_id = request.POST.get('sucursal')
        comprobante_pago = request.FILES.get('comprobante_pago')

        sucursal_instance = sucursal.objects.filter(idSucursal=sucursal_id).first() if sucursal_id else None

        if tipo_entrega == "retiro_tienda":
            pedido = Pedido.objects.create(
                carrito=carrito,
                numero=str(uuid.uuid4()),
                nombre=nombre,
                apellido=apellido,
                direccion=None,
                region=None,
                comuna=None,
                correo=None,
                run=run,
                sucursal=sucursal_instance,
                tipo_entrega=tipo_entrega,
                comprobante_pago=comprobante_pago
            )
        else:
            direccion = request.POST.get('direccion', '')
            region_id = request.POST.get('region')
            comuna_id = request.POST.get('comuna')
            correo = request.POST.get('correo')


            pedido = Pedido.objects.create(
                carrito=carrito,
                numero=str(uuid.uuid4()),
                nombre=nombre,
                apellido=apellido,
                direccion=direccion,
                correo=correo,
                run=None,
                sucursal=sucursal_instance,
                tipo_entrega=tipo_entrega,          
                comprobante_pago=comprobante_pago
            )

        items_carrito = carrito.itemcarrito_set.all()
        for item in items_carrito:
            ItemPedido.objects.create(pedido=pedido, producto=item.producto, cantidad=item.cantidad)

        carrito.productos.clear()

        return JsonResponse({'success': True, 'numero_pedido': pedido.numero})
    else:
        return JsonResponse({'success': False})


# BOLETA
def boleta(request, numero_pedido):
    usuario = request.user
    pedido = Pedido.objects.get(numero=numero_pedido)
    pedidos = Pedido.objects.filter(carrito__usuario=usuario)

    data = {
         'pedido': pedido,
         'pedidos': pedidos,
         'MEDIA_URL': settings.MEDIA_URL,
    }

    return render(request, 'core/boleta.html', data)


# MIS PEDIDOS
@login_required
def mis_pedidos(request):
    usuario = request.user
    pedidos = Pedido.objects.filter(carrito__usuario=usuario)

    data = {
        'pedidos': pedidos,
        'MEDIA_URL': settings.MEDIA_URL,
    }
    return render(request, 'core/mis_pedidos.html', data)


# ADMINISTRACIÓN DE PEDIDOS
@login_required
def administrar_pedidos(request):
    pedidos = Pedido.objects.all()
    form = EstadoPedido()

    if request.method == 'POST':
        form = EstadoPedido(request.POST)
        if form.is_valid():
            pedido_numero = form.cleaned_data['pedido_numero']
            estado = form.cleaned_data['estado']
            pedido = Pedido.objects.get(numero=pedido_numero)
            pedido.estado = estado
            pedido.save()
    
    data = {
        'pedidos': pedidos,
        'form': form,
        'MEDIA_URL': settings.MEDIA_URL,
    }
    return render(request, 'core/administrar_pedidos.html', data)


# CAMBIAR ESTADO
def cambiar_estado(request, numero_orden):
    pedido = Pedido.objects.get(numero=numero_orden)
    form = EstadoPedido(request.POST or None, initial={'estado': pedido.estado.descripcion})

    if request.method == 'POST' and form.is_valid():
        estado_descripcion = form.cleaned_data['estado']
        seguimiento, created = Seguimiento.objects.get_or_create(descripcion=estado_descripcion)
        pedido.estado = seguimiento
        pedido.save()
        return redirect(reverse('administrar_pedidos'))

    data = {
        'pedidos': Pedido.objects.all(),
        'form': form,
    }

    return render(request, 'core/administrar_pedidos.html', data)


# SOPORTE CONTACTO
def soporte_contacto(request):
    return render(request, 'core/soporte_contacto.html')


# PEDIDOS ENTREGADOS — Requiere reemplazo de SP
def pedidos_entregados(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    pedidos = []

    mes_num = int(mes) if mes and mes.isdigit() else 0
    anio_num = int(anio) if anio and anio.isdigit() else 0

    pedidos_qs = Pedido.objects.filter(estado__descripcion='Entregado')

    if mes_num:
        pedidos_qs = pedidos_qs.filter(fecha__month=mes_num)
    if anio_num:
        pedidos_qs = pedidos_qs.filter(fecha__year=anio_num)

    for p in pedidos_qs:
        pedidos.append({
            'id': p.id,
            'numero': p.numero,
            'fecha': p.fecha,
            'carrito_id': p.carrito.id,
            'estado': p.estado.descripcion if p.estado else '',
            'apellido': p.apellido,
            'comuna': p.comuna.nombre if p.comuna else '',
            'correo': p.correo,
            'direccion': p.direccion,
            'nombre': p.nombre,
            'region': p.region.nombre if p.region else '',
            'sucursal_id': p.sucursal.idSucursal if p.sucursal else '',
            'run': p.run,
            'tipo_entrega': p.tipo_entrega,
            'total_pagado': sum([i.producto.precioProducto * i.cantidad for i in p.itempedido_set.all()])
        })

    meses = [str(i).zfill(2) for i in range(1, 13)]
    current_year = datetime.now().year
    years = list(range(2020, current_year + 1))

    data = {
        'pedidos': pedidos,
        'meses': meses,
        'years': years,
        'selected_mes': mes,
        'selected_anio': anio,
    }
    return render(request, 'core/pedidos_entregados.html', data)

# GENERAR INFORMES
def generar_informes(request):
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')

    mes_num = int(mes) if mes and mes.isdigit() else 0
    anio_num = int(anio) if anio and anio.isdigit() else 0

    # Obtener pedidos entregados según filtros
    pedidos_qs = Pedido.objects.filter(estado__descripcion='Entregado')

    if mes_num:
        pedidos_qs = pedidos_qs.filter(fecha__month=mes_num)
    if anio_num:
        pedidos_qs = pedidos_qs.filter(fecha__year=anio_num)

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active

    headers = [
        'Número de Pedido', 'Fecha', 'Cliente', 'Dirección',
        'Comuna', 'Región', 'Estado Actual', 'Total Pagado'
    ]
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Agregar datos de pedidos
    for p in pedidos_qs:
        total_pagado = sum([i.producto.precioProducto * i.cantidad for i in p.itempedido_set.all()])
        ws.append([
            p.numero,
            p.fecha.strftime('%d/%m/%Y') if p.fecha else '',
            f"{p.nombre} {p.apellido}",
            p.direccion or '',
            p.comuna.nombre if p.comuna else '',
            p.region.nombre if p.region else '',
            p.estado.descripcion if p.estado else '',
            f"${total_pagado:,.2f} CLP"
        ])

    # Ajustar columnas
    column_widths = [20, 15, 30, 30, 20, 20, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    filename = f'ventas_entregadas_{mes or "None"}_{anio or "None"}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


# OBTENER SUCURSALES (AJAX)
def obtener_sucursales(request):
    try:
        sucursales = sucursal.objects.all()
        data = [{
            'idSucursal': s.idSucursal,
            'nombreSucursal': s.nombreSucursal,
            'direccionSucursal': s.direccionSucursal
        } for s in sucursales]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# PERFIL DE USUARIO
@login_required
def perfil_usuario(request):
    user = request.user
    if request.method == 'POST':
        form = UsuarioCustomForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente!')
            return redirect('perfil_usuario')
    else:
        form = UsuarioCustomForm(instance=user)
    
    return render(request, 'core/perfil_usuario.html', {'form': form})


# RESET DE CONTRASEÑA (EMAIL)
class CustomPasswordResetView(auth_views.PasswordResetView):
    email_template_name = 'registration/password_reset_email.html'


class PasswordResetView(auth_views.PasswordResetView):
    email_template_name = 'registration/password_reset_email.html'
    subject_template_name = 'registration/password_reset_subject.txt'

